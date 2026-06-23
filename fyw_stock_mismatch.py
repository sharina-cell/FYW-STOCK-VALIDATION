#!/usr/bin/env python3
"""
FYW Stock Mismatch Checker
===========================
Compares the TC Product Master against marketplace stock files
(Shopee, Lazada, TikTok, Zalora) for Melissa, Ipanema, and CSpace,
and produces a formatted Excel report of all mismatched SKUs.

REQUIREMENTS
------------
    pip install pandas openpyxl

Shopee files require LibreOffice installed and on PATH (for .xlsx -> .csv
conversion, due to an openpyxl/Shopee export-format incompatibility):
    - Windows: https://www.libreoffice.org/download/download/
    - Mac:     brew install --cask libreoffice
    - Linux:   sudo apt install libreoffice

HOW TO USE
----------
1. Put this script in a folder together with your files:
     - TC_PRODUCT_MASTER.csv   (or ALL.csv - either filename works)
     - Any of: *_LAZADA.xlsx, *_SHOPEE.xlsx, *_TIKTOK.xlsx, *_ZALORA.xlsx
       (filenames should start with the brand: MELISSA_, IPANEMA_, CSPACE_)
2. Run:
     python fyw_stock_mismatch.py
   Optionally point it at a different folder:
     python fyw_stock_mismatch.py --input-dir "C:\\path\\to\\files" --output "Report.xlsx"
3. Open the generated Stock_Mismatch_Report.xlsx

WHAT IT DOES
------------
- Matches every marketplace file it finds to a brand (from filename prefix)
  and platform (from filename) automatically. Handles common typos/variants
  (e.g. CPSACE_LAZADA.xlsx, trailing underscores, MELISSA_SHOPEE_COMBINED.xlsx).
- Reads stock from the correct column per platform (see COLUMN MAP below).
- Flags any SKU where marketplace stock != TC Master stock.
- Outputs one Excel workbook with:
    - "ALL_MISMATCHES" summary sheet (one row per unique mismatched SKU)
    - One sheet per marketplace file, color-coded:
        RED   = marketplace stock HIGHER than TC Master
        GREEN = marketplace stock LOWER than TC Master

COLUMN MAP (0-indexed positions used when headers are unreliable)
-------------------------------------------------------------------
    TC Product Master : SKU = 'sellerSKU' column, Stock = 'MyStock-Location quantity' column
    Lazada             : SKU = column I (index 8), Stock = column M (index 12); skip first 3 data rows
    Shopee             : SKU = 'SKU' column, Stock = 'Stock' column (after LibreOffice CSV conversion)
    TikTok             : SKU = 'Seller SKU' column, Stock = 'Quantity' column; header on row 3 (index 2)
    Zalora             : SKU = column A (index 0), Stock = column C (index 2)
"""

import argparse
import glob
import os
import re
import shutil
import subprocess
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────

BRANDS = ['MELISSA', 'IPANEMA', 'CSPACE', 'CPSACE']  # CPSACE = common typo for CSPACE
BRAND_DISPLAY = {'MELISSA': 'Melissa', 'IPANEMA': 'Ipanema', 'CSPACE': 'CSpace', 'CPSACE': 'CSpace'}

HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
CELL_FONT   = Font(name='Arial', size=10)
ALT_FILL    = PatternFill('solid', start_color='DDEEFF')
RED_FILL    = PatternFill('solid', start_color='FFE0E0')   # marketplace stock HIGHER than TC
GRN_FILL    = PatternFill('solid', start_color='E0FFE0')   # marketplace stock LOWER than TC
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left', vertical='center')
THIN   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ──────────────────────────────────────────────────────────────────────────
# FILE DISCOVERY
# ──────────────────────────────────────────────────────────────────────────

def find_tc_master(input_dir):
    """Locate the TC Product Master file (handles ALL.csv or TC_PRODUCT_MASTER*.csv)."""
    candidates = (
        glob.glob(os.path.join(input_dir, 'TC_PRODUCT_MASTER*.csv')) +
        glob.glob(os.path.join(input_dir, 'ALL.csv')) +
        glob.glob(os.path.join(input_dir, 'all.csv'))
    )
    if not candidates:
        raise FileNotFoundError(
            "Could not find TC Product Master file. Expected a file named "
            "'TC_PRODUCT_MASTER.csv' or 'ALL.csv' in the input folder."
        )
    return candidates[0]


def detect_brand(filename):
    upper = filename.upper()
    for b in BRANDS:
        if upper.startswith(b):
            return BRAND_DISPLAY[b]
    return None


def detect_platform(filename):
    upper = filename.upper()
    if 'LAZADA' in upper:
        return 'Lazada'
    if 'SHOPEE' in upper:
        return 'Shopee'
    if 'TIKTOK' in upper:
        return 'TikTok'
    if 'ZALORA' in upper:
        return 'Zalora'
    return None


def discover_marketplace_files(input_dir, tc_master_path):
    """Scan input_dir for marketplace files, grouped by (brand, platform)."""
    found = {}  # (brand, platform) -> filepath
    for path in glob.glob(os.path.join(input_dir, '*.xlsx')):
        if os.path.abspath(path) == os.path.abspath(tc_master_path):
            continue
        fname = os.path.basename(path)
        brand = detect_brand(fname)
        platform = detect_platform(fname)
        if brand and platform:
            key = (brand, platform)
            # Prefer files without "_COMBINED" suffix conflicts; just take first found,
            # but if a "COMBINED" variant exists, prefer it (it's usually the merged file).
            if key not in found or 'COMBINED' in fname.upper():
                found[key] = path
    return found


# ──────────────────────────────────────────────────────────────────────────
# LIBREOFFICE CONVERSION (Shopee files only — openpyxl compatibility issue)
# ──────────────────────────────────────────────────────────────────────────

def convert_to_csv_via_libreoffice(xlsx_path, out_dir):
    """Convert a single xlsx to csv using headless LibreOffice. Returns csv path."""
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if not soffice:
        raise RuntimeError(
            "LibreOffice not found on PATH. Shopee files require LibreOffice for "
            "conversion (openpyxl cannot read Shopee's export format directly). "
            "Install LibreOffice and ensure 'soffice' or 'libreoffice' is on your PATH."
        )
    subprocess.run(
        [soffice, '--headless', '--convert-to', 'csv', '--outdir', out_dir, xlsx_path],
        check=True, capture_output=True
    )
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    csv_path = os.path.join(out_dir, f'{base}.csv')
    if not os.path.exists(csv_path):
        raise RuntimeError(f"LibreOffice conversion failed for {xlsx_path}")
    return csv_path


# ──────────────────────────────────────────────────────────────────────────
# PER-PLATFORM READERS  -> each returns DataFrame[Seller_SKU, Market_Stock, Source]
# ──────────────────────────────────────────────────────────────────────────

def read_tiktok(path, brand):
    df = pd.read_excel(path, header=2, dtype=str)
    df = df.iloc[2:].reset_index(drop=True)  # skip "Mandatory"/"Uneditable" rows
    df = df[['Seller SKU', 'Quantity']].rename(
        columns={'Seller SKU': 'Seller_SKU', 'Quantity': 'Market_Stock'})
    df['Seller_SKU'] = df['Seller_SKU'].astype(str).str.strip()
    df['Market_Stock'] = pd.to_numeric(df['Market_Stock'], errors='coerce')
    df['Source'] = f'{brand} - TikTok'
    return df.dropna(subset=['Seller_SKU'])


def read_lazada(path, brand):
    df = pd.read_excel(path, header=0, dtype=str)
    df = df.iloc[3:].reset_index(drop=True)   # skip 3 header/description rows
    df = df.iloc[:, [8, 12]].copy()           # col I = SellerSKU, col M = Quantity
    df.columns = ['Seller_SKU', 'Market_Stock']
    df['Seller_SKU'] = df['Seller_SKU'].astype(str).str.strip()
    df['Market_Stock'] = pd.to_numeric(df['Market_Stock'], errors='coerce')
    df['Source'] = f'{brand} - Lazada'
    return df.dropna(subset=['Seller_SKU'])


def read_shopee(path, brand, tmp_dir):
    """Shopee files need LibreOffice conversion first due to an openpyxl
    incompatibility with Shopee's exported xlsx format."""
    csv_path = convert_to_csv_via_libreoffice(path, tmp_dir)
    # Some Shopee exports have 2 extra system rows before the real header,
    # others have clean headers at row 0. Detect which we have.
    first_cell = str(pd.read_csv(csv_path, nrows=1, header=None).iloc[0, 0])
    if 'et_title' in first_cell or 'sales_info' in first_cell:
        df = pd.read_csv(csv_path, skiprows=2, dtype=str)
        df = df.iloc[2:].reset_index(drop=True)  # skip "Mandatory" + blank row
    else:
        df = pd.read_csv(csv_path, dtype=str)
    df = df[['SKU', 'Stock']].copy()
    df.columns = ['Seller_SKU', 'Market_Stock']
    df['Seller_SKU'] = df['Seller_SKU'].astype(str).str.strip()
    df['Market_Stock'] = pd.to_numeric(df['Market_Stock'], errors='coerce')
    df['Source'] = f'{brand} - Shopee'
    return df.dropna(subset=['Seller_SKU'])


def read_zalora(path, brand):
    df = pd.read_excel(path, dtype=str)       # header row 0 is the real header
    df = df.iloc[:, [0, 2]].copy()            # col A = SellerSku, col C = Quantity
    df.columns = ['Seller_SKU', 'Market_Stock']
    df['Seller_SKU'] = df['Seller_SKU'].astype(str).str.strip()
    df['Market_Stock'] = pd.to_numeric(df['Market_Stock'], errors='coerce')
    df['Source'] = f'{brand} - Zalora'
    return df.dropna(subset=['Seller_SKU'])


READERS = {
    'TikTok': lambda path, brand, tmp: read_tiktok(path, brand),
    'Lazada': lambda path, brand, tmp: read_lazada(path, brand),
    'Shopee': lambda path, brand, tmp: read_shopee(path, brand, tmp),
    'Zalora': lambda path, brand, tmp: read_zalora(path, brand),
}


# ──────────────────────────────────────────────────────────────────────────
# TC MASTER LOADER
# ──────────────────────────────────────────────────────────────────────────

def load_tc_master(path):
    tc = pd.read_csv(path, dtype=str)
    tc['sellerSKU'] = tc['sellerSKU'].astype(str).str.strip()
    tc['MyStock-Location quantity'] = pd.to_numeric(tc['MyStock-Location quantity'], errors='coerce')
    tc = tc.dropna(subset=['sellerSKU'])
    stock_dict = tc.set_index('sellerSKU')['MyStock-Location quantity'].to_dict()
    title_dict = tc.set_index('sellerSKU')['Item title'].to_dict() if 'Item title' in tc.columns else {}
    return stock_dict, title_dict


# ──────────────────────────────────────────────────────────────────────────
# COMPARISON
# ──────────────────────────────────────────────────────────────────────────

def compare_stocks(all_dfs, tc_dict, tc_title):
    mismatch_rows = []
    for df in all_dfs:
        for _, row in df.dropna(subset=['Seller_SKU']).iterrows():
            sku, mkt, source = row['Seller_SKU'], row['Market_Stock'], row['Source']
            if sku in tc_dict:
                tc_val = tc_dict[sku]
                if pd.notna(mkt) and pd.notna(tc_val) and int(mkt) != int(tc_val):
                    mismatch_rows.append({
                        'Seller SKU': sku,
                        'Item Title': tc_title.get(sku, ''),
                        'Source': source,
                        'TC Stock (Master)': int(tc_val),
                        'Marketplace Stock': int(mkt),
                        'Difference (Mktplace - TC)': int(mkt) - int(tc_val),
                    })
    mismatch_df = pd.DataFrame(mismatch_rows)
    if mismatch_df.empty:
        return mismatch_df
    return (mismatch_df
            .drop_duplicates(subset=['Seller SKU', 'Source'])
            .sort_values(['Seller SKU', 'Source'])
            .reset_index(drop=True))


# ──────────────────────────────────────────────────────────────────────────
# EXCEL REPORT BUILDER
# ──────────────────────────────────────────────────────────────────────────

def write_sheet(ws, title_text, headers, rows, col_widths, show_diff_color=True):
    ws.append([title_text])
    ws.row_dimensions[1].height = 22
    tcell = ws.cell(1, 1)
    tcell.font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    tcell.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws.append(headers)
    ws.row_dimensions[2].height = 17
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(2, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for r_idx, row_data in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 16
        diff = row_data[-1] if show_diff_color else 0
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = val
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c_idx <= 2 else CENTER
            if show_diff_color:
                if diff > 0:
                    cell.fill = RED_FILL
                elif diff < 0:
                    cell.fill = GRN_FILL
                elif r_idx % 2 == 1:
                    cell.fill = ALT_FILL
            elif r_idx % 2 == 1:
                cell.fill = ALT_FILL

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(mismatch_df, output_path):
    wb = Workbook()

    ws_sum = wb.create_sheet('ALL_MISMATCHES', 0)
    if mismatch_df.empty:
        write_sheet(ws_sum, 'Stock Mismatch Summary — No mismatches found! 🎉',
                    ['Seller SKU', 'Item Title', 'TC Stock (Master)', 'Affected Marketplaces'],
                    [], [30, 45, 18, 55], show_diff_color=False)
    else:
        sku_sources = (mismatch_df.groupby('Seller SKU')['Source']
                        .apply(lambda x: ', '.join(sorted(x))).reset_index())
        sku_sources.columns = ['Seller SKU', 'Affected Marketplaces']
        sku_base = mismatch_df.drop_duplicates('Seller SKU')[
            ['Seller SKU', 'Item Title', 'TC Stock (Master)']]
        summary = (sku_base.merge(sku_sources, on='Seller SKU')
                   .sort_values('Seller SKU').reset_index(drop=True))
        write_sheet(ws_sum,
                    f'Stock Mismatch Summary — All Marketplaces ({len(summary)} unique SKUs)',
                    ['Seller SKU', 'Item Title', 'TC Stock (Master)', 'Affected Marketplaces'],
                    [tuple(r) for r in summary.itertuples(index=False)],
                    [30, 45, 18, 55], show_diff_color=False)

        for source in sorted(mismatch_df['Source'].unique()):
            safe = re.sub(r'[^A-Za-z0-9_]', '_', source.replace(' - ', '_'))[:31]
            ws = wb.create_sheet(safe)
            df_src = mismatch_df[mismatch_df['Source'] == source].drop(columns=['Source']).reset_index(drop=True)
            write_sheet(ws,
                        f'Stock Mismatch: {source} ({len(df_src)} records)',
                        ['Seller SKU', 'Item Title', 'TC Stock (Master)',
                         'Marketplace Stock', 'Difference (Mktplace - TC)'],
                        [tuple(r) for r in df_src.itertuples(index=False)],
                        [30, 45, 18, 18, 22])

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(output_path)


# ──────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='FYW Stock Mismatch Checker')
    parser.add_argument('--input-dir', default='.', help='Folder containing TC Master + marketplace files (default: current folder)')
    parser.add_argument('--output', default='Stock_Mismatch_Report.xlsx', help='Output Excel filename')
    args = parser.parse_args()

    input_dir = os.path.abspath(args.input_dir)
    print(f"Scanning: {input_dir}")

    tc_path = find_tc_master(input_dir)
    print(f"TC Product Master: {os.path.basename(tc_path)}")
    tc_dict, tc_title = load_tc_master(tc_path)
    print(f"  -> {len(tc_dict)} SKUs loaded")

    files = discover_marketplace_files(input_dir, tc_path)
    if not files:
        print("No marketplace files found (expected *_LAZADA.xlsx, *_SHOPEE.xlsx, "
              "*_TIKTOK.xlsx, or *_ZALORA.xlsx with a MELISSA/IPANEMA/CSPACE prefix).")
        sys.exit(1)

    print(f"Found {len(files)} marketplace file(s):")
    for (brand, platform), path in files.items():
        print(f"  - {brand} {platform}: {os.path.basename(path)}")

    all_dfs = []
    with tempfile.TemporaryDirectory() as tmp_dir:
        for (brand, platform), path in files.items():
            try:
                reader = READERS[platform]
                df = reader(path, brand, tmp_dir)
                all_dfs.append(df)
                print(f"  Loaded {brand} {platform}: {len(df)} rows")
            except Exception as e:
                print(f"  WARNING: failed to read {os.path.basename(path)}: {e}")

        print("\nComparing stocks...")
        mismatch_df = compare_stocks(all_dfs, tc_dict, tc_title)

    if mismatch_df.empty:
        print("No mismatches found across any marketplace! 🎉")
    else:
        print(f"Total mismatch records: {len(mismatch_df)}")
        print(f"Unique mismatched SKUs: {mismatch_df['Seller SKU'].nunique()}")

    out_path = os.path.join(input_dir, args.output)
    build_report(mismatch_df, out_path)
    print(f"\nSaved report: {out_path}")


if __name__ == '__main__':
    main()
