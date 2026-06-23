"""
FYW Stock Mismatch Checker — Streamlit App
==========================================
Upload your TC Product Master (CSV) + any marketplace files (xlsx),
then click Run to get a formatted Excel mismatch report.
"""

import io
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FYW Stock Mismatch Checker",
    page_icon="📦",
    layout="centered",
)

st.title("📦 FYW Stock Mismatch Checker")
st.caption("Compare TC Product Master stock against Shopee, Lazada, TikTok & Zalora listings.")

# ── Styling constants ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="DDEEFF")
RED_FILL    = PatternFill("solid", start_color="FFE0E0")
GRN_FILL    = PatternFill("solid", start_color="E0FFE0")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BRANDS = {
    "MELISSA":  "Melissa",
    "IPANEMA":  "Ipanema",
    "CSPACE":   "CSpace",
    "CPSACE":   "CSpace",   # common typo
}

# ── Helpers ────────────────────────────────────────────────────────────────

def detect_brand(filename: str):
    upper = filename.upper()
    for key, display in BRANDS.items():
        if upper.startswith(key):
            return display
    return None


def detect_platform(filename: str):
    upper = filename.upper()
    if "LAZADA"  in upper: return "Lazada"
    if "SHOPEE"  in upper: return "Shopee"
    if "TIKTOK"  in upper: return "TikTok"
    if "ZALORA"  in upper: return "Zalora"
    return None


# ── File readers ───────────────────────────────────────────────────────────

def read_tc_master(file) -> tuple[dict, dict]:
    tc = pd.read_csv(file, dtype=str)
    tc["sellerSKU"] = tc["sellerSKU"].astype(str).str.strip()
    tc["MyStock-Location quantity"] = pd.to_numeric(
        tc["MyStock-Location quantity"], errors="coerce"
    )
    tc = tc.dropna(subset=["sellerSKU"])
    stock = tc.set_index("sellerSKU")["MyStock-Location quantity"].to_dict()
    title = (
        tc.set_index("sellerSKU")["Item title"].to_dict()
        if "Item title" in tc.columns else {}
    )
    return stock, title


def read_tiktok(file, brand: str) -> pd.DataFrame:
    df = pd.read_excel(file, header=2, dtype=str, engine="calamine")
    df = df.iloc[2:].reset_index(drop=True)   # skip Mandatory / Uneditable rows
    df = df[["Seller SKU", "Quantity"]].rename(
        columns={"Seller SKU": "Seller_SKU", "Quantity": "Market_Stock"}
    )
    df["Seller_SKU"]   = df["Seller_SKU"].astype(str).str.strip()
    df["Market_Stock"] = pd.to_numeric(df["Market_Stock"], errors="coerce")
    df["Source"] = f"{brand} - TikTok"
    return df.dropna(subset=["Seller_SKU"])


def read_lazada(file, brand: str) -> pd.DataFrame:
    df = pd.read_excel(file, header=0, dtype=str, engine="calamine")
    df = df.iloc[3:].reset_index(drop=True)   # skip 3 header/description rows
    df = df.iloc[:, [8, 12]].copy()           # col I = SellerSKU, col M = Quantity
    df.columns = ["Seller_SKU", "Market_Stock"]
    df["Seller_SKU"]   = df["Seller_SKU"].astype(str).str.strip()
    df["Market_Stock"] = pd.to_numeric(df["Market_Stock"], errors="coerce")
    df["Source"] = f"{brand} - Lazada"
    return df.dropna(subset=["Seller_SKU"])


def read_shopee(file, brand: str) -> pd.DataFrame:
    """
    Shopee exports have two formats:
      - With system rows: row 0 = internal header, row 1 = sales_info,
        row 2 = column headers, row 3 = 'Mandatory', row 4+ = data
      - Clean: row 0 = column headers, row 1+ = data
    We use calamine (pure Python) — no LibreOffice required.
    """
    df = pd.read_excel(file, header=2, dtype=str, engine="calamine")
    # If first data row still has 'Mandatory' or NaN in SKU col, skip 3 more rows
    if df.shape[0] > 0:
        first_sku = str(df.iloc[0, 5]) if df.shape[1] > 5 else ""
        skip = 3 if (first_sku in ("nan", "", "Mandatory")) else 0
        df = df.iloc[skip:].reset_index(drop=True)
    df = df.iloc[:, [5, 8]].copy()  # col F = SKU, col I = Stock
    df.columns = ["Seller_SKU", "Market_Stock"]
    df["Seller_SKU"]   = df["Seller_SKU"].astype(str).str.strip()
    df["Market_Stock"] = pd.to_numeric(df["Market_Stock"], errors="coerce")
    df["Source"] = f"{brand} - Shopee"
    return df.dropna(subset=["Seller_SKU"])


def read_zalora(file, brand: str) -> pd.DataFrame:
    df = pd.read_excel(file, dtype=str, engine="calamine")
    df = df.iloc[:, [0, 2]].copy()  # col A = SellerSku, col C = Quantity
    df.columns = ["Seller_SKU", "Market_Stock"]
    df["Seller_SKU"]   = df["Seller_SKU"].astype(str).str.strip()
    df["Market_Stock"] = pd.to_numeric(df["Market_Stock"], errors="coerce")
    df["Source"] = f"{brand} - Zalora"
    return df.dropna(subset=["Seller_SKU"])


READERS = {
    "TikTok": read_tiktok,
    "Lazada":  read_lazada,
    "Shopee":  read_shopee,
    "Zalora":  read_zalora,
}

# ── Comparison ─────────────────────────────────────────────────────────────

def compare_stocks(
    all_dfs: list[pd.DataFrame],
    tc_dict: dict,
    tc_title: dict,
) -> pd.DataFrame:
    rows = []
    for df in all_dfs:
        for _, row in df.dropna(subset=["Seller_SKU"]).iterrows():
            sku, mkt, source = row["Seller_SKU"], row["Market_Stock"], row["Source"]
            if sku in tc_dict:
                tc_val = tc_dict[sku]
                if pd.notna(mkt) and pd.notna(tc_val) and int(mkt) != int(tc_val):
                    rows.append({
                        "Seller SKU":              sku,
                        "Item Title":              tc_title.get(sku, ""),
                        "Source":                  source,
                        "TC Stock (Master)":       int(tc_val),
                        "Marketplace Stock":       int(mkt),
                        "Difference (Mktplace - TC)": int(mkt) - int(tc_val),
                    })
    if not rows:
        return pd.DataFrame()
    return (
        pd.DataFrame(rows)
        .drop_duplicates(subset=["Seller SKU", "Source"])
        .sort_values(["Seller SKU", "Source"])
        .reset_index(drop=True)
    )

# ── Excel builder ──────────────────────────────────────────────────────────

def _write_sheet(ws, title_text, headers, rows, col_widths, show_diff_color=True):
    ws.append([title_text])
    ws.row_dimensions[1].height = 22
    tcell = ws.cell(1, 1)
    tcell.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    tcell.alignment = LEFT
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(headers)
    )

    ws.append(headers)
    ws.row_dimensions[2].height = 17
    for c in range(1, len(headers) + 1):
        cell = ws.cell(2, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for r_idx, row_data in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 16
        diff = row_data[-1] if show_diff_color else 0
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = val
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c_idx <= 2 else CENTER
            if show_diff_color:
                if diff > 0:
                    cell.fill = RED_FILL
                elif diff < 0:
                    cell.fill = GRN_FILL
                elif r_idx % 2 == 1:
                    cell.fill = ALT_FILL
            elif r_idx % 2 == 1:
                cell.fill = ALT_FILL

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(mismatch_df: pd.DataFrame) -> bytes:
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("ALL_MISMATCHES", 0)
    if mismatch_df.empty:
        _write_sheet(
            ws_sum,
            "Stock Mismatch Summary — No mismatches found! 🎉",
            ["Seller SKU", "Item Title", "TC Stock (Master)", "Affected Marketplaces"],
            [],
            [30, 45, 18, 55],
            show_diff_color=False,
        )
    else:
        sku_sources = (
            mismatch_df.groupby("Seller SKU")["Source"]
            .apply(lambda x: ", ".join(sorted(x)))
            .reset_index()
        )
        sku_sources.columns = ["Seller SKU", "Affected Marketplaces"]
        sku_base = mismatch_df.drop_duplicates("Seller SKU")[
            ["Seller SKU", "Item Title", "TC Stock (Master)"]
        ]
        summary = (
            sku_base.merge(sku_sources, on="Seller SKU")
            .sort_values("Seller SKU")
            .reset_index(drop=True)
        )
        _write_sheet(
            ws_sum,
            f"Stock Mismatch Summary — All Marketplaces ({len(summary)} unique SKUs)",
            ["Seller SKU", "Item Title", "TC Stock (Master)", "Affected Marketplaces"],
            [tuple(r) for r in summary.itertuples(index=False)],
            [30, 45, 18, 55],
            show_diff_color=False,
        )

        # ── Per-source sheets ──────────────────────────────────────────────
        for source in sorted(mismatch_df["Source"].unique()):
            safe = re.sub(r"[^A-Za-z0-9_]", "_", source.replace(" - ", "_"))[:31]
            ws = wb.create_sheet(safe)
            df_src = (
                mismatch_df[mismatch_df["Source"] == source]
                .drop(columns=["Source"])
                .reset_index(drop=True)
            )
            _write_sheet(
                ws,
                f"Stock Mismatch: {source} ({len(df_src)} records)",
                [
                    "Seller SKU", "Item Title", "TC Stock (Master)",
                    "Marketplace Stock", "Difference (Mktplace - TC)",
                ],
                [tuple(r) for r in df_src.itertuples(index=False)],
                [30, 45, 18, 18, 22],
            )

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────────

st.markdown("### 1️⃣ Upload TC Product Master")
tc_file = st.file_uploader(
    "TC_PRODUCT_MASTER.csv or ALL.csv",
    type=["csv"],
    key="tc_master",
)

st.markdown("### 2️⃣ Upload Marketplace Files")
st.caption(
    "Name files like `MELISSA_SHOPEE.xlsx`, `IPANEMA_LAZADA.xlsx`, "
    "`CSPACE_TIKTOK.xlsx`, `MELISSA_ZALORA.xlsx` etc. "
    "You can upload as many as you have — not all are required."
)
market_files = st.file_uploader(
    "Marketplace files (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True,
    key="market_files",
)

# Show what's been detected
if market_files:
    st.markdown("**Detected files:**")
    for f in market_files:
        brand    = detect_brand(f.name)
        platform = detect_platform(f.name)
        if brand and platform:
            st.write(f"  ✅ `{f.name}` → **{brand} — {platform}**")
        else:
            st.write(f"  ⚠️ `{f.name}` — could not detect brand/platform (will be skipped)")

st.markdown("---")
run = st.button("🚀 Run Stock Mismatch Check", type="primary", disabled=not (tc_file and market_files))

if run:
    with st.spinner("Loading TC Product Master…"):
        try:
            tc_dict, tc_title = read_tc_master(tc_file)
            st.success(f"TC Master loaded: **{len(tc_dict):,} SKUs**")
        except Exception as e:
            st.error(f"Failed to read TC Master: {e}")
            st.stop()

    all_dfs = []
    seen_keys: dict = {}   # (brand, platform) -> filename, for dedup

    with st.spinner("Reading marketplace files…"):
        for f in market_files:
            fname    = f.name
            brand    = detect_brand(fname)
            platform = detect_platform(fname)
            if not brand or not platform:
                st.warning(f"⚠️ Skipped `{fname}` — cannot detect brand or platform.")
                continue

            key = (brand, platform)
            # Prefer _COMBINED variants; otherwise first-seen wins
            if key in seen_keys:
                if "COMBINED" not in fname.upper():
                    st.info(f"ℹ️ Skipped duplicate `{fname}` (already have {seen_keys[key]})")
                    continue
                else:
                    st.info(f"ℹ️ Replacing `{seen_keys[key]}` with COMBINED file `{fname}`")

            try:
                reader = READERS[platform]
                df = reader(f, brand)
                all_dfs.append(df)
                seen_keys[key] = fname
                st.write(f"  ✅ **{brand} — {platform}**: {len(df):,} rows")
            except Exception as e:
                st.warning(f"⚠️ Could not read `{fname}`: {e}")

    if not all_dfs:
        st.error("No marketplace files could be read. Please check filenames and formats.")
        st.stop()

    with st.spinner("Comparing stocks…"):
        mismatch_df = compare_stocks(all_dfs, tc_dict, tc_title)

    st.markdown("---")
    if mismatch_df.empty:
        st.success("🎉 No mismatches found! All marketplace stocks match the TC Master.")
    else:
        unique_skus = mismatch_df["Seller SKU"].nunique()
        total_recs  = len(mismatch_df)
        st.error(
            f"**{unique_skus:,} unique SKUs** with mismatches "
            f"({total_recs:,} total records across all marketplaces)"
        )

        # Summary table preview
        st.markdown("#### Preview — ALL_MISMATCHES (first 50 rows)")
        sku_sources = (
            mismatch_df.groupby("Seller SKU")["Source"]
            .apply(lambda x: ", ".join(sorted(x)))
            .reset_index()
        )
        sku_sources.columns = ["Seller SKU", "Affected Marketplaces"]
        sku_base = mismatch_df.drop_duplicates("Seller SKU")[
            ["Seller SKU", "Item Title", "TC Stock (Master)"]
        ]
        summary = sku_base.merge(sku_sources, on="Seller SKU").sort_values("Seller SKU")
        st.dataframe(summary.head(50), use_container_width=True, hide_index=True)

        # Per-source breakdown
        st.markdown("#### Breakdown by Marketplace")
        cols = st.columns(3)
        for i, (source, grp) in enumerate(mismatch_df.groupby("Source")):
            cols[i % 3].metric(source, f"{len(grp):,} records")

        st.markdown("---")
        st.markdown("#### 📥 Download Report")
        st.caption(
            "🔴 Red rows = marketplace stock **higher** than TC Master  \n"
            "🟢 Green rows = marketplace stock **lower** than TC Master"
        )

    with st.spinner("Building Excel report…"):
        excel_bytes = build_excel(mismatch_df)

    st.download_button(
        label="⬇️ Download Stock_Mismatch_Report.xlsx",
        data=excel_bytes,
        file_name="Stock_Mismatch_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
